from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from tablas import tabla_rpm_hp, tabla_hueco_acople, tabla_motor, fac_service

# Importar las bibliotecas necesarias y los archivos de tablas

# Definir una función para ingresar la velocidad requerida en RPM y evitar errores en el ingreso.
def RPM_entrada():
    # Repetir hasta que se ingrese un valor numérico válido
    while True:
        try:
            rpm = float(input('Ingresa la velocidad requerida en RPM: '))
            if rpm > 3600:
                print(f'Velocidad de entrada fuera de rango. {rpm} > 3600 rpm\n')
            elif rpm < 5:
                print(f'Velocidad de entrada fuera de rango. {rpm} < 900 rpm\n')
            else:
                break
        except ValueError:
            print('Ingrese un valor numérico\n')
    return rpm

# Función para ingresar la potencia requerida en HP y evitar errores en el ingreso.
def HP_entrada():
    # Repetir hasta que se ingrese un valor numérico válido
    while True:
        try:
            hp = float(input('Ingresa la potencia requerida en hp: '))
            if hp > 68068:
                print(f'Potencia fuera de rango. {hp} > 400 HP\n')
            elif hp < 0.036:
                print(f'Velocidad de entrada fuera de rango. {hp} < 1 HP\n')
            else:
                break
        except ValueError:
            print('Ingrese un valor numérico\n')
    return hp

# Función que busca los acoples en la tabla "tabla_rpm_hp" según los datos ingresados.
def buscar_acople(rpm, hp):
    rpm_tabla = None
    hp_tabla = None
    acople = None
    
    # Recorrer las filas de la tabla hasta encontrar la primera fila con una velocidad RPM inferior a la velocidad requerida
    for fila in tabla_rpm_hp:
        if fila["RPM"] < rpm:
            break
        rpm_tabla = fila["RPM"]

    # Recorrer las filas de la tabla hasta encontrar la primera fila con una velocidad RPM igual a la velocidad de la tabla
    for fila in tabla_rpm_hp:
        if fila["RPM"] == rpm_tabla:
            if fila["HP"] >= hp:
                hp_tabla = fila["HP"]
                break

    # Recorrer las filas de la tabla hasta encontrar la primera fila que coincida con la velocidad RPM y la potencia HP de la tabla
    for fila in tabla_rpm_hp:
        if fila["RPM"] == rpm_tabla and fila["HP"] == hp_tabla:
            acople = fila
    
    # Imprimir información del acople encontrado
    print(f'Rpm_tabla: {rpm_tabla}, hp_equivalente: {hp}, hp_tabla: {hp_tabla}, modelo: {acople["Tamaño"]}')
    return acople

# Función que busca el motor en la tabla "tabla_motor" según los datos ingresados.
def escoge_motor(rpm, hp):
    rpm_tabla = None
    hp_tabla = None
    motor_elegido = None
    
    # Recorrer los motores en la tabla hasta encontrar el primer motor con unavelocidad RPM inferior a la velocidad requerida

    for motor in tabla_motor:
        if motor["RPM"] < rpm:
            break
        rpm_tabla = motor["RPM"]

    # Recorrer los motores en la tabla hasta encontrar el primer motor con una velocidad RPM igual a la velocidad de la tabla
    for motor in tabla_motor:
        if motor["RPM"] == rpm_tabla:
            if motor["HP"] >= hp:
                hp_tabla = motor["HP"]
                break

    # Recorrer los motores en la tabla hasta encontrar el primer motor que coincida con la velocidad RPM y la potencia HP de la tabla
    for motor in tabla_motor:
        if motor["RPM"] == rpm_tabla and motor["HP"] == hp_tabla:
            motor_elegido = motor
    
    # Imprimir información del motor elegido
    print(f'Rpm: {rpm_tabla}, hp: {hp_tabla}, modelo: {motor_elegido["ARMAZON"]}, diámetro: {motor_elegido["U"]}')
    return motor_elegido

# Función que busca las dimensiones del acople en la tabla "tabla_hueco_acople"
def dim(acople):
    dim_acople = None
    pos = 0
    
    # Recorrer las dimensiones de los acoples en la tabla "tabla_hueco_acople" hasta encontrar la primera fila que coincida con el tamaño del acople requerido
    for index, dimensiones in enumerate(tabla_hueco_acople):
        if dimensiones["Tamaño"] == acople["Tamaño"]:
            dim_acople = dimensiones
            pos = index + 1
            break
    
    return dim_acople, pos

# Función para seleccionar otro acople si el acople elegido no cumple con las dimensiones del eje
def reeleccion(pos, dim_eje):
    new_acop = None

    # Recorrer las filas de la tabla "tabla_hueco_acople" a partir de la posición dada
    for i in range(pos, 19):
        new_acop = tabla_hueco_acople[i]["Tamaño"]
        new_dim_min = tabla_hueco_acople[i]["Min"]
        new_dim_max = tabla_hueco_acople[i]["Max"]

        # Comprobar si las dimensiones del eje cumplen con las dimensiones mínimas y máximas del acople en la tabla
        if dim_eje >= new_dim_min and dim_eje <= new_dim_max:
            print("El acople a elegir es:", new_acop, "\n")
            break
        else:
            new_acop = None
    
    if new_acop is None:
        print("No se encontró acople requerido\n")

# Función para guardar los datos en un archivo de Excel
def guardar_datos(acople, motor_elegido):
    # Cargar el archivo de Excel existente o crear uno nuevo
    try:
        wb = load_workbook("datos.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    # Seleccionar la hoja "Resultados" o crear una nueva hoja si no existe
    if "Resultados" in wb.sheetnames:
        ws = wb["Resultados"]
    else:
        ws = wb.create_sheet("Resultados")  

    ws.cell(row=1, column=1, value="RPM Acople")
    ws.cell(row=1, column=2, value="HP Equivalente")
    ws.cell(row=1, column=3, value="HP Tabla")
    ws.cell(row=1, column=4, value="Modelo Acople")
    ws.cell(row=1, column=5, value="RPM Motor")
    ws.cell(row=1, column=6, value="HP Motor")
    ws.cell(row=1, column=7, value="Modelo Motor")
    ws.cell(row=1, column=8, value="Diámetro Motor")

    # Obtener el número de filas actual en la hoja
    num_filas = ws.max_row

    # Escribir los datos en las columnas correspondientes
    ws.cell(row=num_filas + 1, column=1, value=acople["RPM"])
    ws.cell(row=num_filas + 1, column=2, value=hp)
    ws.cell(row=num_filas + 1, column=3, value=acople["HP"])
    ws.cell(row=num_filas + 1, column=4, value=acople["Tamaño"])
    ws.cell(row=num_filas + 1, column=5, value=motor_elegido["RPM"])
    ws.cell(row=num_filas + 1, column=6, value=motor_elegido["HP"])
    ws.cell(row=num_filas + 1, column=7, value=motor_elegido["ARMAZON"])
    ws.cell(row=num_filas + 1, column=8, value=motor_elegido["U"])

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:    
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Guardar los cambios en el archivo de Excel
    wb.save("datos.xlsx")

# Desde aquí será el código principal
###############################################################
####                Inicio del programa.                   ####
###############################################################
if __name__ == '__main__':
    solid = True

    while solid:
        print("""
###############################################################
####                  Ingreso de datos                     ####
###############################################################\n""")

        rpm = RPM_entrada()
        hp = HP_entrada()

        print("""
###############################################################
####               Elija el factor de servicio             ####
###############################################################\n""")
        
        for maqui in fac_service:
            print(maqui["MAQUINA"], ":", maqui["FACTOR"])

        fs = float(input("\nElija un factor de servicio de alguna de las máquinas, o ingrese uno nuevo: "))

        hp = hp * fs

        print("""
###############################################################
####                    Acople elegido                     ####
###############################################################\n""")

        try:
            acople = buscar_acople(rpm, hp)
        except (TypeError, UnboundLocalError):
            print("No se encontraron acoples que satisfagan los datos ingresados, revise los datos y trate de nuevo")

        print("""
###############################################################
####                     Motor elegido                     ####
###############################################################\n""")

        dim_acople, pos = dim(acople)

        motor_elegido = escoge_motor(rpm, hp/fs)
        dim_eje = motor_elegido["U"]

        if dim_eje >= dim_acople["Min"] and dim_eje <= dim_acople["Max"]:
            pass
        else:
            print("""
###############################################################
####    Acople que cumple con las condiciones del motor    ####
###############################################################\n""")
            reeleccion(pos, dim_eje)

        guardar_datos(acople, motor_elegido)

        solid2 = True
        yn = input("¿Deseas realizar otra operación? [Y/N]\n")

        while solid2:
            if yn == "Y" or yn == "y":
                solid2 = False
            elif yn == "N" or yn == "n":
                solid = False
                solid2 = False
            else:
                print("La opción elegida no es correcta")
                yn = input("Vuelva a elegir [Y/N]\n")